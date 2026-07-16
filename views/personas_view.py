import flet as ft
from controllers.persona_controller import PersonaController
from controllers.catalog_controller import CatalogController
from database.db_config import SessionLocal
from database.models import Persona, Atencion
from core.ui_helpers import mostrar_exito, mostrar_snackbar
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import os
from views.components.socioeconomic_dialog import mostrar_ficha_socioeconomica_dialog


def build_personas_view(page: ft.Page, on_new_click=None):
    """Vista de Gestión de Personas — Diálogos via page.overlay (compatible Flet 0.83)."""

    estado = {"mostrar_activos": True}
    datos_actuales = []  # Lista de datos filtrados para Excel
    txt_activos = ft.Ref[ft.Text]()
    txt_inactivos = ft.Ref[ft.Text]()

    # ── Helper: mostrar cualquier AlertDialog via overlay ──────────────────
    def mostrar_dialogo(dlg):
        """Agrega el diálogo al overlay y lo abre. Compatible con Flet 0.83."""
        # Limpiar diálogos previos del overlay
        page.overlay.clear()
        page.overlay.append(dlg)
        dlg.open = True
        page.update()

    def cerrar_dialogo(dlg):
        dlg.open = False
        page.update()

    # ── Estado de Ordenación ────────────────────────────────────────────────
    sort_info = {"index": None, "ascending": True}

    def al_ordenar(col_idx, ascending):
        sort_info["index"] = col_idx
        sort_info["ascending"] = ascending
        tabla.sort_column_index = col_idx
        tabla.sort_ascending = ascending
        cargar_datos(buscador.value)

    # ── Tabla ───────────────────────────────────────────────────────────────
    tabla = ft.DataTable(
        columns=[
            ft.DataColumn(ft.Text("#", weight="bold"), on_sort=lambda e: al_ordenar(e.column_index, e.ascending)),
            ft.DataColumn(ft.Text("Fecha", weight="bold"), on_sort=lambda e: al_ordenar(e.column_index, e.ascending)),
            ft.DataColumn(ft.Text("DNI/Código", weight="bold"), on_sort=lambda e: al_ordenar(e.column_index, e.ascending)),
            ft.DataColumn(ft.Text("Apellidos y Nombres", weight="bold"), on_sort=lambda e: al_ordenar(e.column_index, e.ascending)),
            ft.DataColumn(ft.Text("Cel./Correo", weight="bold")),
            ft.DataColumn(ft.Text("Modalidad / Registro", weight="bold"), on_sort=lambda e: al_ordenar(e.column_index, e.ascending)),
            ft.DataColumn(ft.Text("Tipo/Caso", weight="bold"), on_sort=lambda e: al_ordenar(e.column_index, e.ascending)),
            ft.DataColumn(ft.Text("Facultad/Escuela", weight="bold"), on_sort=lambda e: al_ordenar(e.column_index, e.ascending)),
            ft.DataColumn(ft.Text("Obs./Dirección", weight="bold")),
            ft.DataColumn(ft.Text("Acciones", weight="bold")),
        ],
        column_spacing=18,
        heading_row_color=ft.Colors.with_opacity(0.1, ft.Colors.BLUE_400),
    )
    buscador = ft.TextField(
        hint_text="Buscar por DNI, Nombre o Código...", prefix_icon=ft.Icons.SEARCH,
        width=250, on_change=lambda e: cargar_datos(e.control.value)
    )

    # ── Filtros de Mes, Año y Modalidad ────────────────────────
    _hoy = datetime.now()
    MESES = ["Enero","Febrero","Marzo","Abril","Mayo","Junio",
             "Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"]
             
    # Cargar modalidades para el filtro
    modalidades_cat = CatalogController.get_modalidades()
    
    # Opciones y variables de estado de multi-selección
    opciones_modalidad = [(str(m.id), m.nombre) for m in modalidades_cat]
    opciones_mes = [(str(i+1), MESES[i]) for i in range(12)]
    años_options = [str(a) for a in range(2026, _hoy.year + 5)]
    opciones_año = [(a, a) for a in años_options]
    
    selected_modalidades = set(str(m.id) for m in modalidades_cat) # Todas seleccionadas por defecto
    selected_meses = {str(_hoy.month)} # Mes actual por defecto
    selected_años = {str(_hoy.year)} # Año actual por defecto

    # Funciones auxiliares para el formato de texto de los botones
    def get_modalidades_text():
        if not selected_modalidades:
            return "Ninguna"
        if len(selected_modalidades) == len(modalidades_cat):
            return "Todas"
        names = [m.nombre for m in modalidades_cat if str(m.id) in selected_modalidades]
        text = ", ".join(names)
        if len(text) > 14:
            return f"{len(selected_modalidades)} Sel."
        return text

    def get_meses_text():
        if not selected_meses:
            return "Ninguno"
        if len(selected_meses) == 12:
            return "Todo el Año"
        names = [MESES[int(m)-1] for m in sorted(list(selected_meses), key=int)]
        text = ", ".join(names)
        if len(text) > 12:
            return f"{len(selected_meses)} Meses"
        return text

    def get_años_text():
        if not selected_años:
            return "Ninguno"
        if len(selected_años) == len(años_options):
            return "Todos"
        text = ", ".join(sorted(list(selected_años)))
        if len(text) > 10:
            return f"{len(selected_años)} Años"
        return text

    # Auxiliar para construir botones estilizados de filtros estilo dropdown
    def crear_filtro_boton(label, get_text_func, on_click_func, width):
        text_control = ft.Text(get_text_func(), size=12, overflow=ft.TextOverflow.ELLIPSIS)
        
        def on_hover(e):
            e.control.border = ft.border.all(1, ft.Colors.BLUE_400 if e.data == "true" else ft.Colors.WHITE_38)
            e.control.update()
            
        btn = ft.Container(
            content=ft.Row([
                ft.Column([
                    ft.Text(label, size=9, color=ft.Colors.WHITE_54, weight="bold"),
                    text_control,
                ], spacing=2, alignment=ft.MainAxisAlignment.CENTER, expand=True),
                ft.Icon(ft.Icons.ARROW_DROP_DOWN, color=ft.Colors.WHITE_54, size=20)
            ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN, vertical_alignment="center"),
            width=width,
            height=48,
            border=ft.border.all(1, ft.Colors.WHITE_38),
            border_radius=8,
            padding=ft.padding.symmetric(horizontal=10, vertical=4),
            on_click=on_click_func,
            on_hover=on_hover,
            ink=True,
        )
        
        btn.data = {"update_text": lambda: setattr(text_control, "value", get_text_func())}
        return btn

    # Capa superior de overlay para el selector flotante y barrera transparente
    active_dropdown = {"name": None}
    overlay_layer = ft.Stack(expand=True, visible=False)

    def close_all_dropdowns():
        overlay_layer.controls.clear()
        overlay_layer.visible = False
        active_dropdown["name"] = None
        try: overlay_layer.update()
        except: pass

    # Controlador de menú desplegable con checkmarks que no se cierra al hacer clic
    def toggle_dropdown(name, btn, opciones, selected_set, left_pos, width_val, height_val):
        if active_dropdown["name"] == name:
            close_all_dropdowns()
            return
            
        close_all_dropdowns()
        active_dropdown["name"] = name
        
        # 1. Barrera transparente hit-testable
        barrier = ft.Container(
            expand=True,
            bgcolor=ft.Colors.with_opacity(0.0, ft.Colors.BLACK),
            on_click=lambda _: close_all_dropdowns()
        )
        
        # 2. Contenedor de checkboxes
        checkboxes_col = ft.Column(scroll=ft.ScrollMode.AUTO, tight=True, spacing=5)
        
        def rebuild_dropdown_content():
            checkboxes_col.controls.clear()
            all_selected = len(selected_set) == len(opciones)
            
            def toggle_all(e):
                if all_selected:
                    selected_set.clear()
                else:
                    selected_set.clear()
                    selected_set.update(k for k, _ in opciones)
                btn.data["update_text"]()
                btn.update()
                rebuild_dropdown_content()
                card.update()
                cargar_datos(buscador.value)
                
            chk_all = ft.Checkbox(
                label="Seleccionar Todos",
                value=all_selected,
                on_change=toggle_all,
                fill_color=ft.Colors.BLUE_400
            )
            checkboxes_col.controls.append(chk_all)
            checkboxes_col.controls.append(ft.Divider(height=5, color=ft.Colors.WHITE_24))
            
            for key, label in opciones:
                is_checked = key in selected_set
                
                def make_on_change(k):
                    def on_change(e):
                        if e.control.value:
                            selected_set.add(k)
                        else:
                            selected_set.discard(k)
                        btn.data["update_text"]()
                        btn.update()
                        rebuild_dropdown_content()
                        card.update()
                        cargar_datos(buscador.value)
                    return on_change
                    
                chk = ft.Checkbox(
                    label=label,
                    value=is_checked,
                    on_change=make_on_change(key)
                )
                checkboxes_col.controls.append(chk)
                
        rebuild_dropdown_content()
        
        card = ft.Card(
            content=ft.Container(
                content=checkboxes_col,
                padding=10,
                width=width_val,
                height=height_val,
            ),
            left=left_pos,
            top=48, # Nacimiento exacto en el borde inferior del botón
            elevation=10,
            shadow_color=ft.Colors.BLACK
        )
        
        overlay_layer.controls.append(barrier)
        overlay_layer.controls.append(card)
        overlay_layer.visible = True
        overlay_layer.update()

    # Crear botones de filtro (anchos ajustados para que no desborden la fila horizontal)
    btn_modalidad_filtro = crear_filtro_boton(
        "Modalidad", 
        get_modalidades_text, 
        lambda _: toggle_dropdown("Modalidad", btn_modalidad_filtro, opciones_modalidad, selected_modalidades, 260, 150, 200), 
        150
    )
    btn_mes_filtro = crear_filtro_boton(
        "Mes", 
        get_meses_text, 
        lambda _: toggle_dropdown("Mes", btn_mes_filtro, opciones_mes, selected_meses, 420, 140, 220), 
        120
    )
    btn_año_filtro = crear_filtro_boton(
        "Año", 
        get_años_text, 
        lambda _: toggle_dropdown("Año", btn_año_filtro, opciones_año, selected_años, 550, 140, 180), 
        90
    )

    def limpiar_mes(e):
        selected_meses.clear()
        selected_meses.add(str(datetime.now().month))
        selected_años.clear()
        selected_años.add(str(datetime.now().year))
        selected_modalidades.clear()
        selected_modalidades.update(str(m.id) for m in modalidades_cat)
        buscador.value = ""
        close_all_dropdowns()
        cargar_datos()
        page.update()

    # ── Edición ─────────────────────────────────────────────────────────────
    def abrir_edicion(a_id):
        from database.models import Atencion
        db_s = SessionLocal()
        a = db_s.query(Atencion).filter(Atencion.id == a_id).first()
        if not a:
            db_s.close(); return
        p = db_s.query(Persona).filter(Persona.id == a.persona_id).first()
        if not p:
            db_s.close(); return
            
        tipos = CatalogController.get_tipos_usuario()
        casos = CatalogController.get_casos_sociales()
        facultades = CatalogController.get_facultades()
        modalidades = CatalogController.get_modalidades()

        e_nombres  = ft.TextField(label="Nombres",    value=p.nombres,           expand=1)
        e_apellidos= ft.TextField(label="Apellidos",  value=p.apellidos,          expand=1)
        e_dni      = ft.TextField(label="DNI",        value=p.dni,                expand=1)
        e_codigo   = ft.TextField(label="Código",     value=p.codigo_estudiante or "", expand=1)
        e_edad     = ft.TextField(label="Edad",       value=str(p.edad) if p.edad else "", expand=1)
        e_fecha    = ft.TextField(label="Fecha",      value=a.fecha_atencion.strftime("%d/%m/%Y"), expand=1)
        e_celular  = ft.TextField(label="Celular",    value=p.celular or "",      expand=1)
        e_correo   = ft.TextField(label="Correo",     value=p.correo or "",       expand=1)
        e_año      = ft.Dropdown(
            label="Año Est.", value=p.año_estudio, expand=1,
            options=[ft.dropdown.Option(str(i), f"{i}° Año") for i in range(1, 11)] + [ft.dropdown.Option("Egresado", "Egresado")]
        )
        e_direccion= ft.TextField(label="Dirección",  value=p.direccion or "",    expand=True)
        e_obs      = ft.TextField(label="Observaciones", value=a.observaciones or "", expand=True, multiline=True)
        e_sexo = ft.Dropdown(label="Sexo", value=p.sexo, expand=1,
            options=[ft.dropdown.Option("F","Femenino"), ft.dropdown.Option("M","Masculino")])
        e_tipo = ft.Dropdown(label="Tipo Usuario", value=str(p.tipo_usuario_id) if p.tipo_usuario_id else None, expand=1,
            options=[ft.dropdown.Option(str(t.id), t.nombre) for t in tipos])
        e_caso = ft.Dropdown(label="Caso Social", value=str(a.caso_social_id) if a.caso_social_id else None, expand=1,
            options=[ft.dropdown.Option(str(c.id), c.nombre) for c in casos])
        e_facu = ft.Dropdown(label="Facultad", value=str(p.facultad_id) if p.facultad_id else None, expand=1,
            options=[ft.dropdown.Option(str(f.id), f.nombre) for f in facultades])
        e_escu = ft.Dropdown(label="Escuela", value=str(p.escuela_id) if p.escuela_id else None, expand=1)
        
        # ── Modalidades en Edición ──
        e_mod = ft.Dropdown(label="Modalidad", value=str(a.modalidad_id) if a.modalidad_id else "1", expand=1,
            options=[ft.dropdown.Option(str(m.id), m.nombre) for m in modalidades])
            
        e_reg_mod = ft.TextField(label="Reg. Modalidad", value=a.registro_modalidad or "", expand=1)
        
        def ue_mod(ev):
            from database.db_config import SessionLocal
            from database.models import CatModalidad
            if e_mod.value:
                db = SessionLocal()
                m = db.query(CatModalidad).filter(CatModalidad.id == int(e_mod.value)).first()
                db.close()
                if m and m.nombre not in ["General", "CEPREVAL"]:
                    if m.nombre == "Discapacidad":
                        e_reg_mod.label = "Código CONADIS"
                    else:
                        e_reg_mod.label = f"N° Registro/Carnet {m.nombre} (Opcional)"
                    e_reg_mod.visible = True
                else:
                    e_reg_mod.visible = False
            else:
                e_reg_mod.visible = False
            if ev: e_reg_mod.update()
            
        e_mod.on_select = ue_mod
        
        db_s.close()

        def ue(ev):
            if e_facu.value:
                esc = CatalogController.get_escuelas_by_facultad(int(e_facu.value))
                e_escu.options = [ft.dropdown.Option(str(es.id), es.nombre) for es in esc]
            if ev: e_escu.update()
        e_facu.on_select = ue; ue(None); e_escu.value = str(p.escuela_id) if p.escuela_id else None
        
        # Ejecutar inicialización de visibilidad de modalidad
        ue_mod(None)

        dlg = ft.AlertDialog(modal=True)

        def guardar(e):
            db = SessionLocal()
            at = db.query(Atencion).filter(Atencion.id == a_id).first()
            if at:
                per = db.query(Persona).filter(Persona.id == at.persona_id).first()
                if per:
                    per.nombres = e_nombres.value.upper(); per.apellidos = e_apellidos.value.upper()
                    per.dni = e_dni.value; per.codigo_estudiante = e_codigo.value
                    per.edad = int(e_edad.value) if e_edad.value.isdigit() else None
                    per.sexo = e_sexo.value
                    per.tipo_usuario_id = int(e_tipo.value) if e_tipo.value else None
                    per.facultad_id = int(e_facu.value) if e_facu.value else None
                    per.escuela_id = int(e_escu.value) if e_escu.value else None
                    per.celular = e_celular.value
                    per.correo = e_correo.value; per.direccion = e_direccion.value
                    per.año_estudio = e_año.value
                    
                at.caso_social_id = int(e_caso.value) if e_caso.value else None
                at.modalidad_id = int(e_mod.value) if e_mod.value else None
                at.registro_modalidad = e_reg_mod.value if e_reg_mod.visible else None
                at.observaciones = e_obs.value
                try: at.fecha_atencion = datetime.strptime(e_fecha.value, "%d/%m/%Y")
                except: pass
                db.commit()
            db.close()
            mostrar_exito(page, "✔ Registro actualizado")
            cerrar_dialogo(dlg)
            cargar_datos(buscador.value)

        dlg.title = ft.Text("✏ Editar Registro", weight="bold", size=18)
        dlg.content = ft.Container(width=750, content=ft.Column([
            ft.Row([e_dni, e_fecha, e_edad], spacing=10),
            ft.Row([e_nombres, e_apellidos], spacing=10),
            ft.Row([e_sexo, e_codigo, e_año], spacing=10),
            ft.Row([e_tipo, e_caso], spacing=10),
            ft.Row([e_facu, e_escu], spacing=10),
            ft.Row([e_mod, e_reg_mod], spacing=10),
            ft.Row([e_celular, e_correo], spacing=10),
            e_direccion, e_obs,
        ], spacing=14, scroll=ft.ScrollMode.AUTO, tight=True))
        dlg.actions = [
            ft.TextButton("Cancelar", on_click=lambda _: cerrar_dialogo(dlg)),
            ft.ElevatedButton("Guardar Todo", on_click=guardar,
                bgcolor=ft.Colors.GREEN_800, color=ft.Colors.WHITE, icon=ft.Icons.SAVE),
        ]
        dlg.actions_alignment = "end"
        mostrar_dialogo(dlg)

    # ── Borrado ──────────────────────────────────────────────────────────────
    def abrir_borrado(p_id):
        dlg = ft.AlertDialog(modal=True)
        def confirmar(e):
            PersonaController.eliminar_permanente(p_id)
            mostrar_exito(page, "✔ Registro eliminado")
            cerrar_dialogo(dlg); cargar_datos(buscador.value)
        dlg.title   = ft.Text("⚠ Confirmar Eliminación")
        dlg.content = ft.Text("¿Borrar definitivamente? Esta acción no se puede deshacer.")
        dlg.actions = [
            ft.TextButton("Cancelar", on_click=lambda _: cerrar_dialogo(dlg)),
            ft.ElevatedButton("Sí, Borrar", on_click=confirmar,
                bgcolor=ft.Colors.RED_800, color=ft.Colors.WHITE),
        ]
        dlg.actions_alignment = "end"
        mostrar_dialogo(dlg)

    # ── Estado de Paginación ──────────────────────────────────────────────────
    estado_p = {
        "pagina": 1,
        "por_pagina": 20,
        "total_registros": 0
    }
    
    txt_paginacion = ft.Text("Página 1", size=13, weight="bold")
    btn_prev = ft.IconButton(ft.Icons.NAVIGATE_BEFORE, on_click=lambda _: cambiar_pag(-1), disabled=True)
    btn_next = ft.IconButton(ft.Icons.NAVIGATE_NEXT, on_click=lambda _: cambiar_pag(1))
    
    def on_per_page_change(e):
        estado_p["por_pagina"] = int(e.control.value)
        estado_p["pagina"] = 1
        cargar_datos(buscador.value)

    dd_per_page = ft.Dropdown(
        value="20", width=93, height=45,
        options=[ft.dropdown.Option("10"), ft.dropdown.Option("20"), 
                 ft.dropdown.Option("50"), ft.dropdown.Option("100")],
        on_select=on_per_page_change,
        content_padding=ft.padding.all(10)
    )

    def cambiar_pag(delta):
        estado_p["pagina"] += delta
        cargar_datos(buscador.value)

    # ── Exportación Excel ────────────────────────────────────────────────────
    def exportar_excel(e):
        if not datos_actuales:
            mostrar_snackbar(page, "No hay datos para exportar", "red")
            return
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            headers = ["N°", "DNI", "APELLIDO Y NOMBRES", "EDAD", "SEXO", "TIPO USUARIO", "CODIGO EST.", "ESCUELA", "CASO SOCIAL"]
            fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            font = Font(color="FFFFFF", bold=True)
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.fill = fill; cell.font = font; cell.alignment = Alignment(horizontal="center")
            
            for i, p in enumerate(datos_actuales, 1):
                ws.append([
                    i, 
                    p.get("dni","-"), 
                    f"{p['apellidos']}, {p['nombres']}".upper(), 
                    p.get("edad","-"), 
                    p.get("sexo","-"), 
                    p.get("tipo_usuario","-"), 
                    p.get("codigo_estudiante","-"), 
                    p.get("escuela","-"), 
                    p.get("caso_social","-")
                ])
                
            # Nombre de período
            if len(selected_meses) == 12 or not selected_meses:
                periodo = "Todo_El_Año"
            elif len(selected_meses) == 1:
                periodo = MESES[int(list(selected_meses)[0])-1]
            else:
                names = [MESES[int(m)-1][:3] for m in sorted(list(selected_meses), key=int)]
                periodo = "_".join(names)[:30]

            # Nombre de modalidad
            if len(selected_modalidades) == len(modalidades_cat) or not selected_modalidades:
                mod_nombre = "Todas"
            elif len(selected_modalidades) == 1:
                mod_id = list(selected_modalidades)[0]
                mod_nombre = next((m.nombre for m in modalidades_cat if str(m.id) == mod_id), "Modalidad")
            else:
                names = [m.nombre for m in modalidades_cat if str(m.id) in selected_modalidades]
                mod_nombre = "_".join(names)[:30]

            # Nombre de año
            if len(selected_años) == 1:
                año_nombre = list(selected_años)[0]
            else:
                año_nombre = "_".join(sorted(list(selected_años)))[:20]
                
            estado_tag = "Activos" if estado["mostrar_activos"] else "Inactivos"
            filename = f"Reporte_Estudiantes_{estado_tag}_{mod_nombre}_{periodo}_{año_nombre}.xlsx"
            filepath = os.path.join(os.path.expanduser("~"), "Downloads", filename)
            try: wb.save(filepath); mostrar_exito(page, f"✔ en Descargas")
            except: wb.save(filename); mostrar_exito(page, f"✔ en CarpetaLocal")
        except Exception as ex: mostrar_snackbar(page, f"Error: {str(ex)}", "red")

    # ── Spinner de carga ──────────────────────────────────────────────────────
    spinner_p = ft.Container(
        content=ft.Column([
            ft.ProgressRing(width=50, height=50, stroke_width=4),
            ft.Text("Cargando datos...", size=12, color=ft.Colors.BLUE_200)
        ], horizontal_alignment="center", spacing=12),
        alignment=ft.Alignment(0, 0),
        expand=True,
        visible=False
    )

    # ── Cargar filas ─────────────────────────────────────────────────────────
    def cargar_datos(filtro=""):
        spinner_p.visible = True
        try: spinner_p.update()
        except: pass

        p_all = PersonaController.get_all(solo_activos=False)
        
        # 1. Filtrar por mes (selected_meses)
        if selected_meses:
            p_all = [p for p in p_all if str(p["fecha_atencion"].month) in selected_meses]
        else:
            p_all = []

        # 2. Filtrar por año (selected_años)
        if selected_años:
            p_all = [p for p in p_all if str(p["fecha_atencion"].year) in selected_años]
        else:
            p_all = []
            
        # 3. Filtrar por modalidad (selected_modalidades)
        general_ids = [str(m.id) for m in modalidades_cat if m.nombre == "General"]
        general_id = general_ids[0] if general_ids else None
        
        def match_modalidad(p):
            p_mod_id = p.get("modalidad_id")
            if p_mod_id is None:
                return general_id in selected_modalidades if general_id else True
            return str(p_mod_id) in selected_modalidades
            
        if selected_modalidades:
            p_all = [p for p in p_all if match_modalidad(p)]
        else:
            p_all = []
        
        # 3. Filtrar por pestaña Activo/Inactivo
        p_filtered  = [p for p in p_all if p["activo"] == estado["mostrar_activos"]]
        
        # 4. Filtrar por buscador
        if filtro:
            f = filtro.upper()
            p_filtered = [p for p in p_filtered if
                f in (p["dni"] or "") or
                f in (p["apellidos"] or "").upper() or
                f in (p["nombres"] or "").upper() or
                f in (p["codigo_estudiante"] or "").upper() or
                f in (p["modalidad"] or "").upper()]
        
        # 5. Ordenación Dinámica según cabecera
        if sort_info["index"] is not None:
            idx = sort_info["index"]
            asc = sort_info["ascending"]
            
            if idx == 1: # Fecha
                p_filtered.sort(key=lambda p: p["fecha_atencion"], reverse=not asc)
            elif idx == 2: # DNI
                p_filtered.sort(key=lambda p: (p["dni"] or ""), reverse=not asc)
            elif idx == 3: # Apellidos
                p_filtered.sort(key=lambda p: ((p["apellidos"] or "").upper(), (p["nombres"] or "").upper()), reverse=not asc)
            elif idx == 5: # Modalidad
                p_filtered.sort(key=lambda p: (p["modalidad"] or ""), reverse=not asc)
            elif idx == 6: # Tipo de Usuario
                p_filtered.sort(key=lambda p: (p["tipo_usuario"] or ""), reverse=not asc)
            elif idx == 7: # Facultad/Escuela
                p_filtered.sort(key=lambda p: ((p["facultad"] or "").upper(), (p["escuela"] or "").upper()), reverse=not asc)
            elif idx == 0: # Correlativo
                p_filtered.sort(key=lambda p: p["id"], reverse=not asc)
        else:
            p_filtered.sort(key=lambda p: ((p["apellidos"] or "").upper(), (p["nombres"] or "").upper()))
        
        nonlocal datos_actuales
        datos_actuales = p_filtered

        # 6. Lógica de Paginación (Slicing)
        total = len(p_filtered)
        estado_p["total_registros"] = total
        max_pags = max(1, (total + estado_p["por_pagina"] - 1) // estado_p["por_pagina"])
        
        if estado_p["pagina"] > max_pags: estado_p["pagina"] = max_pags
        if estado_p["pagina"] < 1: estado_p["pagina"] = 1
        
        inicio = (estado_p["pagina"] - 1) * estado_p["por_pagina"]
        fin = inicio + estado_p["por_pagina"]
        
        datos_paginados = p_filtered[inicio:fin]
        
        # Actualizar UI de paginación
        txt_paginacion.value = f"Página {estado_p['pagina']} de {max_pags} ({total} registros)"
        btn_prev.disabled = estado_p["pagina"] <= 1
        btn_next.disabled = estado_p["pagina"] >= max_pags
        
        # Actualizar contadores de pestañas
        if txt_activos.current:
            txt_activos.current.value  = f"Activos ({sum(1 for p in p_all if p.get('activo'))})"
        if txt_inactivos.current:
            txt_inactivos.current.value= f"Inactivos ({sum(1 for p in p_all if not p.get('activo'))})"

        # 7. Dibujar Tabla
        tabla.rows = []
        for i, p in enumerate(datos_paginados, inicio + 1):
            pid = p["id"]
            
            # Modalidad y Registro en Flet
            reg_cod = p.get("registro_modalidad")
            reg_cod_display = f"Reg: {reg_cod}" if reg_cod else "General" if (p.get("modalidad") == "General" or not p.get("modalidad")) else "Sin Código"
            reg_cod_color = ft.Colors.GREEN_400 if reg_cod else ft.Colors.WHITE_54
            
            tabla.rows.append(ft.DataRow(cells=[
                ft.DataCell(ft.Text(str(i), size=13)),
                ft.DataCell(ft.Text(p["fecha_atencion"].strftime("%d/%m/%Y"), size=12)),
                ft.DataCell(ft.Column([
                    ft.Text(p["dni"] or "-", size=12, weight="bold"),
                    ft.Text(p["codigo_estudiante"] or "-", size=12, color=ft.Colors.BLUE_200, weight="bold")
                ], spacing=0)),
                ft.DataCell(ft.Column([
                    ft.Text(f"{p['apellidos']}, {p['nombres']}", weight="bold", size=12),
                    ft.Text(f"{p['edad'] or '-'} años, {p['sexo'] or '-'} ", size=10, color=ft.Colors.WHITE_54)
                ], spacing=0, width=240)),
                ft.DataCell(ft.Column([
                    ft.Text(p["celular"] or "-", size=12),
                    ft.Text(p["correo"] or "-", size=11, color=ft.Colors.BLUE_200)
                ], spacing=0, width=120)),
                # Modalidad / Registro Cell
                ft.DataCell(ft.Column([
                    ft.Text(p.get("modalidad") or "General", size=11, weight="bold", color=ft.Colors.BLUE_300),
                    ft.Text(reg_cod_display, size=10, color=reg_cod_color, weight="bold")
                ], spacing=0, width=150)),
                ft.DataCell(ft.Column([
                    ft.Text(p["tipo_usuario"], size=11),
                    ft.Text(p["caso_social"], size=11, color=ft.Colors.GREEN_400, weight="bold")
                ], spacing=0, width=100)),
                ft.DataCell(ft.Column([
                    ft.Text(p["facultad"], size=11),
                    ft.Text(p["escuela"], size=10, color=ft.Colors.WHITE_54)
                ], spacing=0, width=150)),
                ft.DataCell(ft.Column([
                    ft.Text(p["observaciones"] or "-", size=11, italic=True),
                    ft.Text(p["direccion"] or "-", size=10, color=ft.Colors.WHITE_54)
                ], spacing=0, width=180)),
                ft.DataCell(ft.Row([
                    ft.IconButton(
                        icon=ft.Icons.EDIT_ROUNDED, icon_color=ft.Colors.BLUE_400,
                        icon_size=20, tooltip="Editar",
                        visible=estado["mostrar_activos"],
                        on_click=lambda _, p=pid: abrir_edicion(p)
                    ),
                    ft.IconButton(
                        icon=ft.Icons.BLOCK_ROUNDED, icon_color=ft.Colors.RED_400,
                        icon_size=20, tooltip="Desactivar",
                        visible=estado["mostrar_activos"],
                        on_click=lambda _, p=pid: (PersonaController.desactivar(p), mostrar_exito(page,"Desactivado"), cargar_datos(buscador.value))
                    ),
                    ft.IconButton(
                        icon=ft.Icons.CHECK_CIRCLE_ROUNDED, icon_color=ft.Colors.GREEN_400,
                        icon_size=20, tooltip="Volver de Inactivo",
                        visible=not estado["mostrar_activos"],
                        on_click=lambda _, p=pid: (PersonaController.activar(p), mostrar_exito(page,"Activado"), cargar_datos(buscador.value))
                    ),
                    ft.IconButton(
                        icon=ft.Icons.DELETE_FOREVER_ROUNDED, icon_color=ft.Colors.RED_900,
                        icon_size=20, tooltip="Borrar Definitivo",
                        visible=not estado["mostrar_activos"],
                        on_click=lambda _, p=pid: abrir_borrado(p)
                    ),
                ], spacing=0)),
            ]))
        # Actualizar textos de los botones de filtro
        try:
            for btn in [btn_modalidad_filtro, btn_mes_filtro, btn_año_filtro]:
                btn.data["update_text"]()
                btn.update()
        except: pass

        spinner_p.visible = False
        page.update()

    # ── Tabs ─────────────────────────────────────────────────────────────────
    ia = ft.Container(height=3, bgcolor=ft.Colors.BLUE_400,     border_radius=5)
    ii = ft.Container(height=3, bgcolor=ft.Colors.TRANSPARENT, border_radius=5)

    def ctab(act):
        estado["mostrar_activos"] = act
        ia.bgcolor = ft.Colors.BLUE_400     if act  else ft.Colors.TRANSPARENT
        ii.bgcolor = ft.Colors.BLUE_400     if not act else ft.Colors.TRANSPARENT
        cargar_datos(buscador.value or "")

    cargar_datos()

    area_tabla = ft.Container(
        content=ft.Row(
            [ft.Column([tabla], scroll=ft.ScrollMode.AUTO)],
            scroll=ft.ScrollMode.ALWAYS,
            vertical_alignment=ft.CrossAxisAlignment.START
        ),
        expand=True,
    )
    zona_contenido_p = ft.Stack([area_tabla, spinner_p], expand=True)

    pagination_bar = ft.Container(
        padding=ft.padding.symmetric(horizontal=10, vertical=5),
        bgcolor=ft.Colors.with_opacity(0.05, ft.Colors.BLUE_700),
        border_radius=8,
        content=ft.Row([
            ft.Row([
                ft.Text("Ver:", size=12, color=ft.Colors.WHITE_54),
                dd_per_page
            ], spacing=10),
            ft.VerticalDivider(width=20),
            txt_paginacion,
            ft.Container(expand=True),
            ft.Row([
                btn_prev,
                btn_next
            ], spacing=5)
        ], alignment=ft.MainAxisAlignment.CENTER)
    )

    tabs_row = ft.Row([
        ft.Column([ft.TextButton(content=ft.Row([ft.Icon(ft.Icons.CHECK_CIRCLE_OUTLINE, size=16),
                                                 ft.Text("Activos", ref=txt_activos)]),
                                 on_click=lambda _: (close_all_dropdowns(), ctab(True))), ia], spacing=2),
        ft.Column([ft.TextButton(content=ft.Row([ft.Icon(ft.Icons.BLOCK_OUTLINED, size=16),
                                                 ft.Text("Inactivos", ref=txt_inactivos)]),
                                 on_click=lambda _: (close_all_dropdowns(), ctab(False))), ii], spacing=2),
    ], spacing=10)

    filters_row = ft.Row([
        buscador,
        btn_modalidad_filtro,
        btn_mes_filtro,
        btn_año_filtro,
        ft.IconButton(ft.Icons.RESTART_ALT_ROUNDED, tooltip="Limpiar Filtros", on_click=limpiar_mes),
    ], spacing=10, wrap=False, vertical_alignment=ft.CrossAxisAlignment.CENTER)

    filters_and_table_stack = ft.Stack([
        ft.Column([
            filters_row,
            zona_contenido_p,
            pagination_bar
        ], expand=True, spacing=12),
        overlay_layer
    ], expand=True)

    main_column = ft.Column([
        ft.Row([
            ft.Icon(ft.Icons.SCHOOL_ROUNDED, color=ft.Colors.GREEN_400, size=28),
            ft.Text("Gestión de Estudiantes", size=24, weight="bold"),
            ft.Container(expand=True),
            ft.ElevatedButton("Excel", icon=ft.Icons.FILE_DOWNLOAD, bgcolor=ft.Colors.GREEN_800, color=ft.Colors.WHITE, on_click=exportar_excel),
            ft.ElevatedButton("+ Nuevo", bgcolor=ft.Colors.GREEN_800, color=ft.Colors.WHITE,
                               on_click=lambda _: (close_all_dropdowns(), on_new_click())),
        ], spacing=10),
        ft.Divider(color=ft.Colors.BLUE_900),
        tabs_row,
        ft.Divider(height=5, color="transparent"),
        filters_and_table_stack
    ], expand=True, spacing=12)

    return ft.Container(
        padding=ft.padding.only(left=25, right=25, top=20, bottom=10),
        expand=True,
        content=main_column
    )
