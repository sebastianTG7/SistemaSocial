import flet as ft
from datetime import datetime
from controllers.persona_controller import PersonaController
from controllers.catalog_controller import CatalogController
from database.db_config import SessionLocal
from database.models import Persona
from core.ui_helpers import mostrar_snackbar, mostrar_exito
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import os


def build_historial_view(page: ft.Page):
    """Vista de Historial Unificado: Informe Mensual, Modalidades, Paginación, Excel y Acciones."""
    
    # ── Estado de Datos y Paginación ──────────────────────────────────────────
    datos_actuales = [] # Lista completa filtrada para Excel
    estado_p = {
        "pagina": 1,
        "por_pagina": 20,
        "total_registros": 0
    }
    
    # ── Helper: mostrar cualquier AlertDialog via overlay ──────────────────
    def mostrar_dialogo(dlg):
        """Agrega el diálogo al overlay y lo abre. Compatible con Flet 0.83."""
        page.overlay.clear()
        page.overlay.append(dlg)
        dlg.open = True
        page.update()

    def cerrar_dialogo(dlg):
        dlg.open = False
        page.update()
        
    # ── Edición ─────────────────────────────────────────────────────────────
    def abrir_edicion(p_id):
        db_s = SessionLocal()
        p = db_s.query(Persona).filter(Persona.id == p_id).first()
        if not p:
            db_s.close(); return
        tipos = CatalogController.get_tipos_usuario()
        casos = CatalogController.get_casos_sociales()
        facultades = CatalogController.get_facultades()
        modalidades = CatalogController.get_modalidades()

        e_nombres  = ft.TextField(label="Nombres",    value=p.nombres,           expand=1)
        e_apellidos= ft.TextField(label="Apellidos",  value=p.apellidos,          expand=1)
        e_dni      = ft.TextField(label="DNI",        value=p.dni,                expand=1)
        e_codigo   = ft.TextField(label="Código",     value=p.codigo_estudiante or "", expand=1)
        e_edad     = ft.TextField(label="Edad",       value=str(p.edad) if p.edad else "", expand=1)
        e_fecha    = ft.TextField(label="Fecha",      value=p.fecha_atencion.strftime("%d/%m/%Y"), expand=1)
        e_celular  = ft.TextField(label="Celular",    value=p.celular or "",      expand=1)
        e_correo   = ft.TextField(label="Correo",     value=p.correo or "",       expand=1)
        e_año      = ft.Dropdown(
            label="Año Est.", value=p.año_estudio, expand=1,
            options=[ft.dropdown.Option(str(i), f"{i}° Año") for i in range(1, 11)] + [ft.dropdown.Option("Egresado", "Egresado")]
        )
        e_direccion= ft.TextField(label="Dirección",  value=p.direccion or "",    expand=True)
        e_obs      = ft.TextField(label="Observaciones", value=p.observaciones or "", expand=True, multiline=True)
        e_sexo = ft.Dropdown(label="Sexo", value=p.sexo, expand=1,
            options=[ft.dropdown.Option("F","Femenino"), ft.dropdown.Option("M","Masculino")])
        e_tipo = ft.Dropdown(label="Tipo Usuario", value=str(p.tipo_usuario_id) if p.tipo_usuario_id else None, expand=1,
            options=[ft.dropdown.Option(str(t.id), t.nombre) for t in tipos])
        e_caso = ft.Dropdown(label="Caso Social", value=str(p.caso_social_id) if p.caso_social_id else None, expand=1,
            options=[ft.dropdown.Option(str(c.id), c.nombre) for c in casos])
        e_facu = ft.Dropdown(label="Facultad", value=str(p.facultad_id) if p.facultad_id else None, expand=1,
            options=[ft.dropdown.Option(str(f.id), f.nombre) for f in facultades])
        e_escu = ft.Dropdown(label="Escuela", value=str(p.escuela_id) if p.escuela_id else None, expand=1)
        
        # ── Modalidades en Edición ──
        e_mod = ft.Dropdown(label="Modalidad", value=str(p.modalidad_id) if p.modalidad_id else "1", expand=1,
            options=[ft.dropdown.Option(str(m.id), m.nombre) for m in modalidades])
            
        e_reg_mod = ft.TextField(label="Reg. Modalidad", value=p.registro_modalidad or "", expand=1)
        
        def ue_mod(ev):
            from database.db_config import SessionLocal
            from database.models import CatModalidad
            if e_mod.value:
                db = SessionLocal()
                m = db.query(CatModalidad).filter(CatModalidad.id == int(e_mod.value)).first()
                db.close()
                if m and m.nombre not in ["General", "CEPREVAL"]:
                    e_reg_mod.label = f"N° Registro/Carnet {m.nombre} (Opcional)"
                    e_reg_mod.visible = True
                else:
                    e_reg_mod.visible = False
            else:
                e_reg_mod.visible = False
            if ev: e_reg_mod.update()
            
        e_mod.on_select = ue_mod
        
        db_s.close()

        def ue(ev):
            if e_facu.value:
                esc = CatalogController.get_escuelas_by_facultad(int(e_facu.value))
                e_escu.options = [ft.dropdown.Option(str(es.id), es.nombre) for es in esc]
            if ev: e_escu.update()
        e_facu.on_select = ue; ue(None); e_escu.value = str(p.escuela_id) if p.escuela_id else None
        
        ue_mod(None)

        dlg = ft.AlertDialog(modal=True)

        def guardar(e):
            db = SessionLocal()
            r = db.query(Persona).filter(Persona.id == p_id).first()
            if r:
                r.nombres = e_nombres.value.upper(); r.apellidos = e_apellidos.value.upper()
                r.dni = e_dni.value; r.codigo_estudiante = e_codigo.value
                r.edad = int(e_edad.value) if e_edad.value.isdigit() else None
                r.sexo = e_sexo.value
                r.tipo_usuario_id = int(e_tipo.value) if e_tipo.value else None
                r.caso_social_id = int(e_caso.value) if e_caso.value else None
                r.facultad_id = int(e_facu.value) if e_facu.value else None
                r.escuela_id = int(e_escu.value) if e_escu.value else None
                r.modalidad_id = int(e_mod.value) if e_mod.value else None
                r.registro_modalidad = e_reg_mod.value if e_reg_mod.visible else None
                r.celular = e_celular.value
                r.correo = e_correo.value; r.direccion = e_direccion.value
                r.año_estudio = e_año.value; r.observaciones = e_obs.value
                try: r.fecha_atencion = datetime.strptime(e_fecha.value, "%d/%m/%Y")
                except: pass
                db.commit()
            db.close()
            mostrar_exito(page, "✔ Registro actualizado")
            cerrar_dialogo(dlg)
            cargar_datos()

        dlg.title = ft.Text("✏ Editar Registro", weight="bold", size=18)
        dlg.content = ft.Container(width=750, content=ft.Column([
            ft.Row([e_dni, e_fecha, e_edad], spacing=10),
            ft.Row([e_nombres, e_apellidos], spacing=10),
            ft.Row([e_sexo, e_codigo, e_año], spacing=10),
            ft.Row([e_tipo, e_caso], spacing=10),
            ft.Row([e_facu, e_escu], spacing=10),
            ft.Row([e_mod, e_reg_mod], spacing=10),
            ft.Row([e_celular, e_correo], spacing=10),
            e_direccion, e_obs,
        ], spacing=14, scroll=ft.ScrollMode.AUTO, tight=True))
        dlg.actions = [
            ft.TextButton("Cancelar", on_click=lambda _: cerrar_dialogo(dlg)),
            ft.ElevatedButton("Guardar Todo", on_click=guardar,
                bgcolor=ft.Colors.GREEN_800, color=ft.Colors.WHITE, icon=ft.Icons.SAVE),
        ]
        dlg.actions_alignment = "end"
        mostrar_dialogo(dlg)

    # ── Estado de Ordenación ────────────────────────────────────────────────
    sort_info = {"index": None, "ascending": True}

    def al_ordenar(col_idx, ascending):
        sort_info["index"] = col_idx
        sort_info["ascending"] = ascending
        tabla.sort_column_index = col_idx
        tabla.sort_ascending = ascending
        cargar_datos()

    # ── Elementos de UI ──────────────────────────────────────────────────────
    tabla = ft.DataTable(
        columns=[
            ft.DataColumn(ft.Text("#", weight="bold"), on_sort=lambda e: al_ordenar(e.column_index, e.ascending)),
            ft.DataColumn(ft.Text("Fecha", weight="bold"), on_sort=lambda e: al_ordenar(e.column_index, e.ascending)),
            ft.DataColumn(ft.Text("DNI/Código", weight="bold"), on_sort=lambda e: al_ordenar(e.column_index, e.ascending)),
            ft.DataColumn(ft.Text("Apellidos y Nombres", weight="bold"), on_sort=lambda e: al_ordenar(e.column_index, e.ascending)),
            ft.DataColumn(ft.Text("Cel./Correo", weight="bold")),
            ft.DataColumn(ft.Text("Modalidad / Registro", weight="bold"), on_sort=lambda e: al_ordenar(e.column_index, e.ascending)),
            ft.DataColumn(ft.Text("Tipo/Caso", weight="bold"), on_sort=lambda e: al_ordenar(e.column_index, e.ascending)),
            ft.DataColumn(ft.Text("Facultad/Escuela", weight="bold"), on_sort=lambda e: al_ordenar(e.column_index, e.ascending)),
            ft.DataColumn(ft.Text("Obs./Dirección", weight="bold")),
        ],
        column_spacing=18,
        heading_row_color=ft.Colors.with_opacity(0.1, ft.Colors.BLUE_400),
    )

    _hoy = datetime.now()
    meses_nombres = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
    
    # Cargar modalidades para el filtro
    modalidades_cat = CatalogController.get_modalidades()
    
    dd_modalidad_filtro = ft.Dropdown(
        label="Modalidad", width=220, value="all",
        options=[
            ft.dropdown.Option(key="all", text="Todas las Modalidades")
        ] + [ft.dropdown.Option(key=str(m.id), text=m.nombre) for m in modalidades_cat],
        on_select=lambda _: reset_pag_y_cargar()
    )

    dd_mes = ft.Dropdown(
        label="Mes", width=155, value=str(_hoy.month),
        options=[
            ft.dropdown.Option(key="all", text="Todo el Año")
        ] + [ft.dropdown.Option(key=str(i+1), text=meses_nombres[i]) for i in range(12)],
        on_select=lambda _: reset_pag_y_cargar()
    )
    
    dd_año = ft.Dropdown(
        label="Año", width=110, value=str(_hoy.year),
        options=[ft.dropdown.Option(key="all", text="Todos")] + [ft.dropdown.Option(key=str(a), text=str(a)) for a in range(2026, _hoy.year + 5)],
        on_select=lambda _: reset_pag_y_cargar()
    )

    buscador = ft.TextField(
        label="Buscar por DNI, Nombre o Código...", prefix_icon=ft.Icons.SEARCH, expand=True, 
        on_change=lambda _: reset_pag_y_cargar()
    )

    txt_paginacion = ft.Text("Página 1", size=13, weight="bold")
    btn_prev = ft.IconButton(ft.Icons.NAVIGATE_BEFORE, on_click=lambda _: cambiar_pag(-1), disabled=True)
    btn_next = ft.IconButton(ft.Icons.NAVIGATE_NEXT, on_click=lambda _: cambiar_pag(1))
    
    dd_per_page = ft.Dropdown(
        value="20", width=93, height=45,
        options=[ft.dropdown.Option("10"), ft.dropdown.Option("20"), 
                 ft.dropdown.Option("50"), ft.dropdown.Option("100")],
        on_select=lambda e: (estado_p.update({"por_pagina": int(e.control.value)}), reset_pag_y_cargar()),
        content_padding=ft.Padding(10,2,10,2)   
    )

    def reset_pag_y_cargar():
        estado_p["pagina"] = 1
        cargar_datos()

    def cambiar_pag(delta):
        estado_p["pagina"] += delta
        cargar_datos()

    # ── Exportación Excel ────────────────────────────────────────────────────
    def exportar_excel(e):
        if not datos_actuales:
            mostrar_snackbar(page, "No hay datos para exportar", "red")
            return
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            headers = ["N°", "DNI", "APELLIDOS Y NOMBRES", "EDAD", "SEXO", "TIPO DE USUARIO", "CODIGO EST", "AÑO DE ESTUDIOS", "FACULTAD", "ESCUELA", "CASO SOCIAL", "MODALIDAD DE INGRESO", "REGISTRO/CARNET MODALIDAD", "OBSERVACIONES"]
            fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            font = Font(color="FFFFFF", bold=True)
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.fill = fill; cell.font = font; cell.alignment = Alignment(horizontal="center")
            
            for i, p in enumerate(datos_actuales, 1):
                ws.append([
                    i, 
                    p.get("dni","-"), 
                    f"{p['apellidos']}, {p['nombres']}".upper(), 
                    p.get("edad","-"), 
                    p.get("sexo","-"), 
                    p.get("tipo_usuario","-"), 
                    p.get("codigo_estudiante","-"), 
                    p.get("año_estudio","-"), 
                    p.get("facultad","-"), 
                    p.get("escuela","-"), 
                    p.get("caso_social","-"),
                    p.get("modalidad","General"),
                    p.get("registro_modalidad","-") or "-",
                    p.get("observaciones","-") or "-"
                ])
                
            periodo = "Anual" if dd_mes.value == "all" else meses_nombres[int(dd_mes.value)-1]
            mod_nombre = "Todas"
            if dd_modalidad_filtro.value != "all" and dd_modalidad_filtro.value is not None:
                mod_nombre = next((m.nombre for m in modalidades_cat if str(m.id) == dd_modalidad_filtro.value), "Modalidad")
                
            filename = f"Informe_Historial_{mod_nombre}_{periodo}_{dd_año.value}.xlsx"
            filepath = os.path.join(os.path.expanduser("~"), "Downloads", filename)
            try: wb.save(filepath); mostrar_exito(page, f"✔ en Descargas")
            except: wb.save(filename); mostrar_exito(page, f"✔ en CarpetaLocal")
        except Exception as ex: mostrar_snackbar(page, f"Error: {str(ex)}", "red")

    # ── Spinner de carga ──────────────────────────────────────────────────────
    spinner = ft.Container(
        content=ft.Column([
            ft.ProgressRing(width=50, height=50, stroke_width=4),
            ft.Text("Cargando datos...", size=12, color=ft.Colors.BLUE_200)
        ], horizontal_alignment="center", spacing=12),
        alignment=ft.Alignment(0, 0),
        expand=True,
        visible=False
    )

    def cargar_datos():
        nonlocal datos_actuales
        spinner.visible = True
        try: spinner.update()
        except: pass
        
        p_all = PersonaController.get_all(solo_activos=False)
        p_all = [p for p in p_all if p["activo"]]
        
        # Filtros de Período
        if dd_mes.value and dd_mes.value != "all": 
            p_all = [p for p in p_all if p["fecha_atencion"].month == int(dd_mes.value)]
        if dd_año.value and dd_año.value != "all": 
            p_all = [p for p in p_all if p["fecha_atencion"].year == int(dd_año.value)]
            
        # Filtro de Modalidad
        filtro_mod = dd_modalidad_filtro.value
        if filtro_mod and filtro_mod != "all":
            p_all = [p for p in p_all if str(p.get("modalidad_id")) == filtro_mod]
            
        # Buscador por texto
        f = buscador.value.upper() if buscador.value else ""
        if f: 
            p_all = [p for p in p_all if
                f in (p["dni"] or "") or
                f in (p["apellidos"] or "").upper() or
                f in (p["nombres"] or "").upper() or
                f in (p["codigo_estudiante"] or "").upper() or
                f in (p["modalidad"] or "").upper()]

        # Ordenación Dinámica según cabecera
        if sort_info["index"] is not None:
            idx = sort_info["index"]
            asc = sort_info["ascending"]
            if idx == 1: # Fecha
                p_all.sort(key=lambda p: p["fecha_atencion"], reverse=not asc)
            elif idx == 2: # DNI
                p_all.sort(key=lambda p: (p["dni"] or ""), reverse=not asc)
            elif idx == 3: # Apellidos
                p_all.sort(key=lambda p: ((p["apellidos"] or "").upper(), (p["nombres"] or "").upper()), reverse=not asc)
            elif idx == 5: # Modalidad
                p_all.sort(key=lambda p: (p["modalidad"] or ""), reverse=not asc)
            elif idx == 6: # Tipo de Usuario
                p_all.sort(key=lambda p: (p["tipo_usuario"] or ""), reverse=not asc)
            elif idx == 7: # Facultad/Escuela
                p_all.sort(key=lambda p: ((p["facultad"] or "").upper(), (p["escuela"] or "").upper()), reverse=not asc)
            elif idx == 0: # Correlativo
                p_all.sort(key=lambda p: p["id"], reverse=not asc)
        else:
            p_all.sort(key=lambda p: ((p["apellidos"] or "").upper(), (p["nombres"] or "").upper()))
            
        datos_actuales = p_all # Guardar para Excel
        
        total = len(p_all)
        ppp = int(dd_per_page.value) if dd_per_page.value else 20
        estado_p["por_pagina"] = ppp

        max_pags = max(1, (total + ppp - 1) // ppp)
        if estado_p["pagina"] > max_pags: estado_p["pagina"] = max_pags
        
        inicio = (estado_p["pagina"] - 1) * ppp
        datos_paginados = p_all[inicio:inicio+ppp]
        
        txt_paginacion.value = f"Página {estado_p['pagina']} de {max_pags} ({total} registros)"
        btn_prev.disabled = estado_p["pagina"] <= 1
        btn_next.disabled = estado_p["pagina"] >= max_pags

        tabla.rows = []
        for i, p in enumerate(datos_paginados, inicio + 1):
            pid = p["id"]
            
            # Modalidad y Registro en Flet
            reg_cod = p.get("registro_modalidad")
            reg_cod_display = f"Reg: {reg_cod}" if reg_cod else "General" if (p.get("modalidad") == "General" or not p.get("modalidad")) else "Sin Código"
            reg_cod_color = ft.Colors.GREEN_400 if reg_cod else ft.Colors.WHITE_54
            
            tabla.rows.append(ft.DataRow(cells=[
                ft.DataCell(ft.Text(str(i), size=13)),
                ft.DataCell(ft.Text(p["fecha_atencion"].strftime("%d/%m/%Y"), size=12)),
                ft.DataCell(ft.Column([ft.Text(p["dni"] or "-", size=12, weight="bold"), ft.Text(p["codigo_estudiante"] or "-", size=12, color=ft.Colors.BLUE_200, weight="bold")], spacing=0)),
                ft.DataCell(ft.Column([ft.Text(f"{p['apellidos']}, {p['nombres']}", weight="bold", size=12), ft.Text(f"{p['edad'] or '-'} años, {p['sexo'] or '-'} ", size=10, color=ft.Colors.WHITE_54)], spacing=0, width=240)),
                ft.DataCell(ft.Column([ft.Text(p["celular"] or "-", size=12), ft.Text(p["correo"] or "-", size=11, color=ft.Colors.BLUE_200)], spacing=0, width=120)),
                ft.DataCell(ft.Column([ft.Text(p.get("modalidad") or "General", size=11, weight="bold", color=ft.Colors.BLUE_300), ft.Text(reg_cod_display, size=10, color=reg_cod_color, weight="bold")], spacing=0, width=150)),
                ft.DataCell(ft.Column([ft.Text(p["tipo_usuario"], size=11), ft.Text(p["caso_social"], size=11, color=ft.Colors.GREEN_400, weight="bold")], spacing=0, width=100)),
                ft.DataCell(ft.Column([ft.Text(p["facultad"], size=11), ft.Text(p["escuela"], size=10, color=ft.Colors.WHITE_54)], spacing=0, width=150)),
                ft.DataCell(ft.Column([ft.Text(p["observaciones"] or "-", size=11, italic=True), ft.Text(p["direccion"] or "-", size=10, color=ft.Colors.WHITE_54)], spacing=0, width=180)),
            ]))
        spinner.visible = False
        page.update()

    cargar_datos()

    # ── Layout Flexible ──────────────────────────────────────────────
    area_tabla = ft.Container(
        content=ft.Row(
            [ft.Column([tabla], scroll=ft.ScrollMode.AUTO)],
            scroll=ft.ScrollMode.ALWAYS,
            vertical_alignment=ft.CrossAxisAlignment.START
        ),
        expand=True,
    )
    zona_contenido = ft.Stack([area_tabla, spinner], expand=True)

    return ft.Container(
        padding=ft.padding.only(left=25, right=25, top=20, bottom=10), expand=True,
        content=ft.Column([
            ft.Row([
                ft.Icon(ft.Icons.HISTORY_ROUNDED, color=ft.Colors.BLUE_400, size=28), ft.Text("Historial de Atenciones Unificado", size=24, weight="bold"),
                ft.Container(expand=True),
                ft.ElevatedButton("Excel", icon=ft.Icons.FILE_DOWNLOAD, bgcolor=ft.Colors.GREEN_800, color=ft.Colors.WHITE, on_click=exportar_excel)
            ]),
            ft.Divider(color=ft.Colors.BLUE_900),
            ft.Row([buscador, dd_modalidad_filtro, dd_mes, dd_año, ft.IconButton(ft.Icons.RESTART_ALT_ROUNDED, on_click=lambda _: reset_pag_y_cargar())], spacing=10),
            
            # Tabla / Spinner
            zona_contenido,
            
            # Barra de Paginacion fija abajo
            ft.Container(
                padding=ft.padding.symmetric(horizontal=10, vertical=5), bgcolor=ft.Colors.with_opacity(0.05, ft.Colors.BLUE_700), border_radius=8,
                content=ft.Row([
                    ft.Row([ft.Text("Ver:", size=11, color=ft.Colors.WHITE_54), dd_per_page], spacing=10),
                    ft.VerticalDivider(width=20),
                    txt_paginacion,
                    ft.Container(expand=True),
                    ft.Row([btn_prev, btn_next], spacing=5)
                ], alignment=ft.MainAxisAlignment.CENTER)
            )
        ], expand=True, spacing=12)
    )
