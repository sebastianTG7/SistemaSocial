import flet as ft
from controllers.persona_controller import PersonaController
from controllers.catalog_controller import CatalogController
from core.ui_helpers import mostrar_exito, mostrar_snackbar
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import os

def build_usuarios_view(page: ft.Page, on_new_click=None):
    """Vista de Gestión de Usuarios (Directorio)."""
    
    estado = {"mostrar_activos": True}
    datos_actuales = []
    
    # ── Helpers Diálogos ──────────────────────────────────────────────────────
    def mostrar_dialogo(dlg):
        for c in list(page.overlay):
            if isinstance(c, ft.AlertDialog):
                page.overlay.remove(c)
        page.overlay.append(dlg)
        dlg.open = True
        page.update()

    def cerrar_dialogo(dlg):
        dlg.open = False
        page.update()

    # ── Catálogos y Filtros ──────────────────────────────────────────────────
    tipos_usuario = CatalogController.get_tipos_usuario()
    
    dd_filtro_tipo = ft.Dropdown(
        label="Tipo de Usuario", width=200,
        options=[ft.dropdown.Option("ALL", "Todos")] + [ft.dropdown.Option(str(t.id), t.nombre) for t in tipos_usuario],
        value="ALL",
        on_select=lambda e: cargar_datos(buscador.value)
    )
    
    from database.db_config import SessionLocal
    from database.models import CatEscuela
    db = SessionLocal()
    escuelas_all = db.query(CatEscuela).filter(CatEscuela.activo == True).order_by(CatEscuela.nombre).all()
    db.close()
    
    dd_filtro_escuela = ft.Dropdown(
        label="Escuela Profesional", width=250,
        options=[ft.dropdown.Option("ALL", "Todas")] + [ft.dropdown.Option(str(e.id), e.nombre) for e in escuelas_all],
        value="ALL",
        on_select=lambda e: cargar_datos(buscador.value)
    )
    
    dd_filtro_año = ft.Dropdown(
        label="Año de Estudio", width=150,
        options=[ft.dropdown.Option("ALL", "Todos")] + [ft.dropdown.Option(str(i), f"{i}° Año") for i in range(1, 11)] + [ft.dropdown.Option("Egresado", "Egresado")],
        value="ALL",
        on_select=lambda e: cargar_datos(buscador.value)
    )

    # ── Tabla ───────────────────────────────────────────────────────────────
    sort_info = {"index": None, "ascending": True}

    def al_ordenar(col_idx, ascending):
        sort_info["index"] = col_idx
        sort_info["ascending"] = ascending
        tabla.sort_column_index = col_idx
        tabla.sort_ascending = ascending
        cargar_datos(buscador.value)

    tabla = ft.DataTable(
        columns=[
            ft.DataColumn(ft.Text("ID", weight="bold"), on_sort=lambda e: al_ordenar(e.column_index, e.ascending)),
            ft.DataColumn(ft.Text("DNI", weight="bold"), on_sort=lambda e: al_ordenar(e.column_index, e.ascending)),
            ft.DataColumn(ft.Text("Apellidos y Nombres", weight="bold"), on_sort=lambda e: al_ordenar(e.column_index, e.ascending)),
            ft.DataColumn(ft.Text("Tipo", weight="bold")),
            ft.DataColumn(ft.Text("Escuela", weight="bold")),
            ft.DataColumn(ft.Text("Año / Cód.", weight="bold")),
            ft.DataColumn(ft.Text("Celular", weight="bold")),
            ft.DataColumn(ft.Text("Acciones", weight="bold")),
        ],
        column_spacing=20,
        heading_row_color=ft.Colors.with_opacity(0.1, ft.Colors.BLUE_400),
    )
    
    buscador = ft.TextField(
        hint_text="Buscar por DNI o Nombres...", prefix_icon=ft.Icons.SEARCH,
        width=300, on_change=lambda e: cargar_datos(e.control.value)
    )

    # ── Paginación ──────────────────────────────────────────────────────────
    items_por_pagina = 15
    pagina_actual = ft.Ref[ft.Text]()
    btn_prev = ft.Ref[ft.IconButton]()
    btn_next = ft.Ref[ft.IconButton]()
    estado_paginacion = {"pagina": 1, "total_paginas": 1}

    def cambiar_pagina(delta):
        estado_paginacion["pagina"] += delta
        renderizar_tabla()

    pagination_bar = ft.Container(
        padding=ft.padding.symmetric(horizontal=10, vertical=5),
        bgcolor=ft.Colors.with_opacity(0.05, ft.Colors.BLUE_700),
        border_radius=8,
        content=ft.Row([
            ft.Text(ref=pagina_actual, value="Página 1 de 1"),
            ft.Container(expand=True),
            ft.Row([
                ft.IconButton(ref=btn_prev, icon=ft.Icons.CHEVRON_LEFT, on_click=lambda _: cambiar_pagina(-1)),
                ft.IconButton(ref=btn_next, icon=ft.Icons.CHEVRON_RIGHT, on_click=lambda _: cambiar_pagina(1)),
            ], spacing=5)
        ], alignment=ft.MainAxisAlignment.CENTER)
    )

    # ── Acciones de Fila (Editar, Desactivar, Eliminar) ──────────────────────
    def confirmar_eliminar_permanente(p_id, nombres):
        dlg = ft.AlertDialog(modal=True)
        def al_si(e):
            PersonaController.eliminar_permanente_persona(p_id)
            cerrar_dialogo(dlg)
            mostrar_exito(page, f"Usuario {nombres} eliminado de forma permanente.")
            cargar_datos(buscador.value)
            
        def al_no(e):
            cerrar_dialogo(dlg)
            
        dlg.title = ft.Text("Confirmar Eliminacion Permanente")
        dlg.content = ft.Text(f"ATENCION: Esta accion eliminara permanentemente a {nombres} de la base de datos.\n\nAl hacer esto, tambien se borraran TODAS sus atenciones, fichas socioeconomicas y derivaciones asociadas (borrado en cascada).\n\n¿Esta completamente seguro de proceder?")
        dlg.actions = [
            ft.TextButton("Cancelar", on_click=al_no),
            ft.ElevatedButton("Si, Eliminar Permanentemente", on_click=al_si, bgcolor=ft.Colors.RED_800, color=ft.Colors.WHITE)
        ]
        dlg.actions_alignment = "end"
        mostrar_dialogo(dlg)

    def desactivar_activar_fila(p_id, nombres, activo):
        if activo:
            PersonaController.desactivar_persona(p_id)
            mostrar_exito(page, f"Usuario {nombres} desactivado.")
        else:
            PersonaController.activar_persona(p_id)
            mostrar_exito(page, f"Usuario {nombres} reactivado.")
        cargar_datos(buscador.value)

    def abrir_modal_editar(p):
        dlg = ft.AlertDialog(modal=True, title=ft.Text("Editar Usuario"))
        
        f_nombres = ft.TextField(label="Nombres", value=p["nombres"], expand=True)
        f_apellidos = ft.TextField(label="Apellidos", value=p["apellidos"], expand=True)
        f_edad = ft.TextField(label="Edad", value=str(p["edad"]) if p["edad"] else "", width=100)
        
        dd_sexo = ft.Dropdown(label="Sexo", width=120, value=p["sexo"], options=[ft.dropdown.Option("F", "Femenino"), ft.dropdown.Option("M", "Masculino")])
        f_codigo = ft.TextField(label="Codigo Estudiante", value=p["codigo_estudiante"] or "", width=180)
        f_celular = ft.TextField(label="Celular", value=p["celular"] or "", width=150)
        f_correo = ft.TextField(label="Correo", value=p["correo"] or "", expand=True)
        f_direccion = ft.TextField(label="Direccion", value=p["direccion"] or "", expand=True)
        
        t_id = str(p.get("tipo_usuario_id") or "")
        e_id = str(p.get("escuela_id") or "")
        
        dd_tipo = ft.Dropdown(
            label="Tipo Usuario", width=200, 
            options=[ft.dropdown.Option(str(t.id), t.nombre) for t in tipos_usuario],
        )
        tipo_match = next((t for t in tipos_usuario if t.nombre == p["tipo_usuario"]), None)
        if tipo_match: dd_tipo.value = str(tipo_match.id)
        
        dd_escuela = ft.Dropdown(
            label="Escuela", width=250,
            options=[ft.dropdown.Option(str(e.id), e.nombre) for e in escuelas_all]
        )
        escuela_match = next((e for e in escuelas_all if e.nombre == p["escuela"]), None)
        if escuela_match: dd_escuela.value = str(escuela_match.id)
        
        dd_año = ft.Dropdown(
            label="Año de Estudio", width=150,
            options=[ft.dropdown.Option(str(i), f"{i}° Año") for i in range(1, 11)] + [ft.dropdown.Option("Egresado", "Egresado")]
        )
        if p["año_estudio"]: dd_año.value = p["año_estudio"]
        
        def guardar_cambios(e):
            datos = {
                "nombres": f_nombres.value,
                "apellidos": f_apellidos.value,
                "edad": f_edad.value,
                "sexo": dd_sexo.value,
                "codigo_estudiante": f_codigo.value,
                "año_estudio": dd_año.value,
                "celular": f_celular.value,
                "correo": f_correo.value,
                "direccion": f_direccion.value,
                "tipo_usuario_id": dd_tipo.value,
                "escuela_id": dd_escuela.value
            }
            exito, msj = PersonaController.actualizar_persona(p["id"], datos)
            if exito:
                mostrar_exito(page, "Datos de usuario actualizados correctamente.")
                cerrar_dialogo(dlg)
                cargar_datos(buscador.value)
            else:
                mostrar_snackbar(page, f"Error: {msj}", "red")

        dlg.content = ft.Container(
            width=650,
            content=ft.Column([
                ft.Row([f_nombres, f_apellidos]),
                ft.Row([f_edad, dd_sexo, dd_tipo]),
                ft.Row([dd_escuela, dd_año, f_codigo]),
                ft.Row([f_celular, f_correo]),
                f_direccion
            ], tight=True, scroll=ft.ScrollMode.AUTO)
        )
        dlg.actions = [
            ft.TextButton("Cancelar", on_click=lambda _: cerrar_dialogo(dlg)),
            ft.ElevatedButton("Guardar", on_click=guardar_cambios, bgcolor=ft.Colors.GREEN_800, color=ft.Colors.WHITE)
        ]
        mostrar_dialogo(dlg)

    # ── Carga y Renderizado ─────────────────────────────────────────────────
    def cargar_datos(termino=""):
        todos = PersonaController.get_all_personas(solo_activos=estado["mostrar_activos"])
        filtrados = []
        term = termino.lower()
        
        f_tipo = dd_filtro_tipo.value
        f_escuela = dd_filtro_escuela.value
        f_año = dd_filtro_año.value
        
        for r in todos:
            # Buscar en DNI o Nombres/Apellidos
            match_busqueda = not term or term in str(r["dni"]).lower() or term in (str(r["nombres"])+" "+str(r["apellidos"])).lower()
            
            # Filtros desplegables
            match_tipo = f_tipo == "ALL" or str(r.get("tipo_usuario_id", "")) == f_tipo
            match_esc = f_escuela == "ALL" or str(r.get("escuela_id", "")) == f_escuela
            match_año = f_año == "ALL" or str(r.get("año_estudio", "")) == f_año
            
            if match_busqueda and match_tipo and match_esc and match_año:
                filtrados.append(r)
                
        # Ordenamiento
        idx = sort_info["index"]
        asc = sort_info["ascending"]
        if idx is not None:
            def sort_key(x):
                if idx == 0: return x["id"] or 0
                if idx == 1: return x["dni"] or ""
                if idx == 2: return f"{x['apellidos']} {x['nombres']}"
                return ""
            filtrados.sort(key=sort_key, reverse=not asc)
            
        nonlocal datos_actuales
        datos_actuales = filtrados
        estado_paginacion["pagina"] = 1
        estado_paginacion["total_paginas"] = max(1, (len(filtrados) + items_por_pagina - 1) // items_por_pagina)
        
        renderizar_tabla()

    def renderizar_tabla():
        if not btn_prev.current: return
        p = estado_paginacion["pagina"]
        tp = estado_paginacion["total_paginas"]
        btn_prev.current.disabled = (p == 1)
        btn_next.current.disabled = (p == tp)
        pagina_actual.current.value = f"Página {p} de {tp} (Total: {len(datos_actuales)})"
        
        inicio = (p - 1) * items_por_pagina
        fin = inicio + items_por_pagina
        pagina_datos = datos_actuales[inicio:fin]
        
        filas = []
        for r in pagina_datos:
            es_activo = r["activo"]
            nombres_completos = f"{r['apellidos']}, {r['nombres']}"
            
            acciones = []
            # Editar
            acciones.append(
                ft.IconButton(
                    icon=ft.Icons.EDIT_ROUNDED, icon_color=ft.Colors.BLUE_400,
                    tooltip="Editar", on_click=lambda e, p=r: abrir_modal_editar(p)
                )
            )
            # Desactivar / Activar
            if es_activo:
                acciones.append(
                    ft.IconButton(
                        icon=ft.Icons.PERSON_OFF_ROUNDED, icon_color=ft.Colors.ORANGE_400,
                        tooltip="Desactivar", on_click=lambda e, id=r["id"], nom=nombres_completos: desactivar_activar_fila(id, nom, True)
                    )
                )
            else:
                acciones.append(
                    ft.IconButton(
                        icon=ft.Icons.PERSON_ADD_ROUNDED, icon_color=ft.Colors.GREEN_400,
                        tooltip="Reactivar", on_click=lambda e, id=r["id"], nom=nombres_completos: desactivar_activar_fila(id, nom, False)
                    )
                )
                acciones.append(
                    ft.IconButton(
                        icon=ft.Icons.DELETE_FOREVER_ROUNDED, icon_color=ft.Colors.RED_700,
                        tooltip="Eliminar Permanentemente", on_click=lambda e, id=r["id"], nom=nombres_completos: confirmar_eliminar_permanente(id, nom)
                    )
                )
            
            color_texto = ft.Colors.WHITE if es_activo else ft.Colors.RED_200
            
            filas.append(ft.DataRow(
                cells=[
                    ft.DataCell(ft.Text(str(r["id"]), color=color_texto)),
                    ft.DataCell(ft.Text(r["dni"], color=color_texto)),
                    ft.DataCell(ft.Text(nombres_completos, color=color_texto)),
                    ft.DataCell(ft.Text(r["tipo_usuario"] or "-", color=color_texto)),
                    ft.DataCell(ft.Text(r["escuela"] or "-", color=color_texto)),
                    ft.DataCell(ft.Text(f"{r['año_estudio'] or '-'} / {r['codigo_estudiante'] or '-'}", color=color_texto)),
                    ft.DataCell(ft.Text(r["celular"] or "-", color=color_texto)),
                    ft.DataCell(ft.Row(acciones, spacing=0))
                ]
            ))
            
        tabla.rows = filas
        page.update()

    def exportar_excel(e):
        if not datos_actuales:
            mostrar_snackbar(page, "No hay datos para exportar", "red")
            return
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            headers = ["ID", "DNI", "APELLIDOS Y NOMBRES", "EDAD", "SEXO", "TIPO USUARIO", "ESCUELA", "AÑO DE ESTUDIO", "CÓDIGO", "CELULAR", "CORREO", "DIRECCIÓN"]
            fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            font = Font(color="FFFFFF", bold=True)
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.fill = fill
                cell.font = font
                cell.alignment = Alignment(horizontal="center")
            
            for i, p in enumerate(datos_actuales, 1):
                ws.append([
                    p.get("id", ""),
                    p.get("dni", ""),
                    f"{p['apellidos']}, {p['nombres']}".upper(),
                    p.get("edad", ""),
                    p.get("sexo", ""),
                    p.get("tipo_usuario", ""),
                    p.get("escuela", ""),
                    p.get("año_estudio", ""),
                    p.get("codigo_estudiante", ""),
                    p.get("celular", ""),
                    p.get("correo", ""),
                    p.get("direccion", "")
                ])
                
            estado_tag = "Activos" if estado["mostrar_activos"] else "Inactivos"
            filename = f"Usuarios_filtros_{estado_tag}.xlsx"
            filepath = os.path.join(os.path.expanduser("~"), "Downloads", filename)
            try:
                wb.save(filepath)
                mostrar_exito(page, f"✔ Excel guardado en Descargas: {filename}")
            except:
                wb.save(filename)
                mostrar_exito(page, f"✔ Excel guardado en Carpeta Local: {filename}")
        except Exception as ex:
            mostrar_snackbar(page, f"Error: {str(ex)}", "red")

    # ── Pestañas Activos/Inactivos ──────────────────────────────────────────
    def ctab(es_activo):
        estado["mostrar_activos"] = es_activo
        cargar_datos(buscador.value)

    tabs = ft.Row([
        ft.TextButton("Usuarios Activos", icon=ft.Icons.CHECK_CIRCLE_OUTLINE, on_click=lambda _: ctab(True)),
        ft.TextButton("Usuarios Inactivos", icon=ft.Icons.BLOCK_OUTLINED, on_click=lambda _: ctab(False)),
    ], spacing=10)

    # ── UI Principal ────────────────────────────────────────────────────────
    area_tabla = ft.Container(
        content=ft.Row(
            [ft.Column([tabla], scroll=ft.ScrollMode.AUTO)],
            scroll=ft.ScrollMode.ALWAYS,
            vertical_alignment=ft.CrossAxisAlignment.START
        ),
        expand=True,
    )
    zona_contenido_p = ft.Stack([area_tabla], expand=True)

    filters_row = ft.Row([
        buscador,
        dd_filtro_tipo,
        dd_filtro_escuela,
        dd_filtro_año,
        ft.IconButton(
            icon=ft.Icons.REFRESH, tooltip="Recargar", 
            on_click=lambda _: cargar_datos(buscador.value)
        )
    ], spacing=15, vertical_alignment=ft.CrossAxisAlignment.CENTER)

    main_column = ft.Column([
        ft.Row([
            ft.Icon(ft.Icons.PEOPLE_ALT_ROUNDED, color=ft.Colors.BLUE_400, size=28),
            ft.Text("Gestión de Usuarios", size=24, weight="bold"),
            ft.Container(expand=True),
            ft.ElevatedButton("Excel", icon=ft.Icons.FILE_DOWNLOAD, bgcolor=ft.Colors.GREEN_800, color=ft.Colors.WHITE, on_click=exportar_excel),
            ft.ElevatedButton("+ Nueva Atención", bgcolor=ft.Colors.GREEN_800, color=ft.Colors.WHITE,
                               on_click=lambda _: on_new_click() if on_new_click else None),
        ], spacing=10),
        ft.Divider(color=ft.Colors.BLUE_900),
        tabs,
        filters_row,
        zona_contenido_p,
        pagination_bar
    ], expand=True, spacing=15)

    # Cargar datos iniciales
    cargar_datos()

    return ft.Container(
        padding=ft.padding.only(left=25, right=25, top=20, bottom=10),
        expand=True,
        content=main_column
    )
